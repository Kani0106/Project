from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect, render
from django.http import HttpResponseRedirect
from django.contrib import messages
from .threads import *
from .models import *
from .utils import *
import xlrd

context = {}

# home page
def home_page(request):
    return render(request, "main/home.html")

# about page
def about_page(request):
    messages.info(request, 'User does not exists. Please Signup')
    return render(request, "main/about.html")


# student login
def student_login(request):
    try:
        if request.method == 'POST':
            email = request.POST.get('email')
            password = request.POST.get('password')
            student_obj = StudentModel.objects.filter(email=email).first()
            if student_obj is None:
                messages.info(request, 'User does not exists. Please Signup')
                return redirect('student-login')
            user = authenticate(email=email, password=password)
            if user is  None:
                messages.info(request, 'Incorrect Password.')
                return redirect('student-login/')
            login(request, user)
            messages.success(request, 'Student Successfully logged in')
            return redirect('student-dashboard')
    except Exception as e:
        print(e)
    return render(request, "accounts/student-login.html")

    
# tpo login
def tpo_login(request):
    try:
        if request.method == 'POST':
            email = request.POST.get('email')
            password = request.POST.get('password')
            teacher_obj = TeacherModel.objects.filter(email=email).first()
            if teacher_obj is None:
                messages.info(request, 'User does not exists.')
                return redirect('signup')  
            user = authenticate(email=email, password=password)
            if user is  None:
                messages.info(request, 'Incorrect Password.')
                return redirect('login')
            login(request, user)
            messages.success(request, 'Successfully logged in')
            return redirect('tpo-dashboard')
    except Exception as e:
        print(e)
    return render(request, "accounts/tpo-login.html")


# student profile
@login_required(login_url='/student-login/')
def student_profile(request):
    context["user"] = StudentModel.objects.get(email=request.user)
    return render(request, "student/profile.html", context)


# student profile
@login_required(login_url='/student-login/')
def update_student_profile(request):
    try:
        user = StudentModel.objects.get(email=request.user)
        if request.method == 'POST':
            user.resume = request.FILES['resume']
            user.headshot = request.POST.get("headshot")
            user.linkedIn_link = request.POST.get("linkedIn_link")
            user.gitHub_link = request.POST.get("gitHub_link")
            user.bio = request.POST.get("bio")
            user.skills = request.POST.get("skills")
            user.save()
            print("@@@@@@@@@")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER')) 
    except Exception as e:
        print(e)
    return render(request, "student/profile.html", context)

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import StudentModel, FileSavingModel
from openpyxl import load_workbook

@login_required(login_url='/tpo-login/')
def add_student_data(request):
    if request.method == 'POST':
        if 'file' in request.FILES:
            try:
                file_obj = FileSavingModel.objects.create(file=request.FILES['file'])
                path = file_obj.file.path
                workbook = load_workbook(filename=path)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Columns mapped to Excel columns
                    name, email, phone, roll_no, password, bio, skills, resume, headshot, linkedIn_link, gitHub_link = row

                    student = StudentModel(
                        name=name,
                        email=email.lower(),
                        phone=phone,
                        roll_no=roll_no,
                        bio=bio,
                        skills=skills,
                        linkedIn_link=linkedIn_link,
                        gitHub_link=gitHub_link
                    )
                    student.set_password(password)
                    if resume:
                        student.resume.save(name, resume)
                    if headshot:
                        student.headshot.save(name, headshot)
                    student.save()
                
                messages.success(request, "File uploaded and students added successfully!")
            except Exception as e:
                messages.error(request, f"Failed to upload file. Error: {e}")
        
        else:
            # Handle direct form submission
            name = request.POST.get('name')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            roll_no = request.POST.get('roll_no')
            password = request.POST.get('password')
            bio = request.POST.get('bio', '')
            skills = request.POST.get('skills', '')
            resume = request.FILES.get('resume')
            headshot = request.FILES.get('headshot')
            linkedIn_link = request.POST.get('linkedIn_link', '')
            gitHub_link = request.POST.get('gitHub_link', '')

            try:
                student = StudentModel(
                    name=name,
                    email=email.lower(),
                    phone=phone,
                    roll_no=roll_no,
                    bio=bio,
                    skills=skills,
                    linkedIn_link=linkedIn_link,
                    gitHub_link=gitHub_link
                )
                student.set_password(password)
                if resume:
                    student.resume = resume
                if headshot:
                    student.headshot = headshot
                student.save()
                
                messages.success(request, "Student added successfully!")
            except Exception as e:
                messages.error(request, f"Error creating student: {e}")
        
        return redirect('add-student-details')  # Redirect after POST to prevent re-submissions

    return render(request, "tpo/add-students.html")


# student logout
@login_required(login_url='/student-login/')
def student_logout(request):
    logout(request)
    messages.info(request, 'Logged Out')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


# tpo logout
@login_required(login_url='/tpo-login/')
def tpo_logout(request):
    logout(request)
    messages.info(request, 'Logged Out')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

# add tpo
def create_tpo(request):
    obj = TeacherModel.objects.create(name="Placement", email="placement@gmail.com")
    obj.set_password("password")
    obj.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER')) 
